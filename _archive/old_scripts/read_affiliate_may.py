"""
Script อ่านข้อมูล Affiliate พ.ค. 2026 จาก Google Sheet (xlsx ที่ download ไว้)
รันใน Terminal: python3 read_affiliate_may.py
"""
import json, base64, io, glob, os

# หาไฟล์ xlsx ที่ download ไว้ล่าสุดใน temp folder
search_patterns = [
    "/var/folders/rp/gcgj8vnn68n9v7flmfth963c0000gn/T/claude-hostloop-plugins/**/*download_file_content*.txt",
    os.path.expanduser("~/Library/Application Support/Claude/**/tool-results/*download_file_content*.txt"),
]

target_file = "/var/folders/rp/gcgj8vnn68n9v7flmfth963c0000gn/T/claude-hostloop-plugins/a8d001833b717686/projects/-Users-thanasablilutanon-Library-Application-Support-Claude-local-agent-mode-sessions-63474075-287d-46b7-bbd8-5f5bc16621e4-b7d79abc-f657-493b-8a54-c51cdb92110f-local-2777b6f1-4477-4091-8dc7-b92b1ddc2a-h44hdp/8fb8b59a-ee7c-4d57-8a3f-532c7d59b8b9/tool-results/mcp-631cf5b2-d348-4eff-a244-06b07fea56f7-download_file_content-1779590249873.txt"

print("กำลังอ่านไฟล์...")
with open(target_file, 'r') as f:
    data = json.load(f)

b64_content = data['content']
print(f"Base64 size: {len(b64_content):,} chars")

xlsx_bytes = base64.b64decode(b64_content)
print(f"XLSX size: {len(xlsx_bytes):,} bytes")

try:
    import openpyxl
except ImportError:
    print("กำลัง install openpyxl...")
    os.system("pip3 install openpyxl --break-system-packages -q")
    import openpyxl

wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
print(f"\nSheet ทั้งหมด: {wb.sheetnames}\n")

# หา sheet พ.ค.
may_sheet = None
for name in wb.sheetnames:
    if 'พค' in name or 'พ.ค' in name:
        may_sheet = wb[name]
        print(f"✅ พบ sheet: {name}")
        break

if not may_sheet:
    print("❌ ไม่พบ sheet พ.ค. — ใช้ sheet ล่าสุด:", wb.sheetnames[-1])
    may_sheet = wb[wb.sheetnames[-1]]

gmv_total = net_total = comm_total = creator_count = 0

for row in may_sheet.iter_rows(min_row=2, values_only=True):
    b = row[1] if len(row) > 1 else None   # GMV
    d = row[3] if len(row) > 3 else None   # Net
    e = row[4] if len(row) > 4 else None   # Commission
    if isinstance(b, (int, float)) and b > 0:
        gmv_total += b
        creator_count += 1
    if isinstance(d, (int, float)):
        net_total += d
    if isinstance(e, (int, float)):
        comm_total += e

print(f"\n{'='*40}")
print(f"📊 ผลลัพธ์ Affiliate พ.ค. 2026")
print(f"{'='*40}")
print(f"GMV รวม     : ฿{gmv_total:,.2f}")
print(f"Net รวม     : ฿{net_total:,.2f}")
print(f"Commission  : ฿{comm_total:,.2f}")
print(f"จำนวน Creator: {creator_count} คน")
print(f"{'='*40}")
