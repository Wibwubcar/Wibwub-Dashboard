#!/usr/bin/env python3
"""
WIBWUB Auto-Update Script (LaunchAgent version)
รันทุกวัน 09:00 / 18:00 ผ่าน macOS LaunchAgent
ถ้า Mac พับตอนนั้น launchd จะรันทันทีตอน Mac ตื่น

ติดตั้ง dependencies (ครั้งแรกครั้งเดียว):
  pip3 install gspread google-auth pandas openpyxl --break-system-packages

ตั้ง Google Service Account (ดูคำแนะนำใน README ด้านล่าง):
  ~/wibwub-service-account.json

=== README: ตั้ง Google Service Account ===
1. ไปที่ https://console.cloud.google.com/
2. สร้าง Project ใหม่ชื่อ "WIBWUB"
3. เปิด "Google Sheets API" (APIs & Services → Enable APIs)
4. IAM & Admin → Service Accounts → Create Service Account
   - ชื่อ: wibwub-updater
   - Role: ไม่ต้องเลือก (กด Continue → Done)
5. คลิก Service Account ที่สร้าง → Keys → Add Key → JSON → Download
6. เปลี่ยนชื่อไฟล์ที่ดาวน์โหลดเป็น wibwub-service-account.json
7. ย้ายไปที่ home folder: mv ~/Downloads/wibwub-service-account.json ~/
8. Copy อีเมล service account (ดูได้ในไฟล์ JSON ที่ "client_email")
9. เปิด Google Sheet ทั้ง 4 ไฟล์ → Share → วางอีเมล service account → Viewer
   - Shopee: https://docs.google.com/spreadsheets/d/10LrzWB8bbCO9FigCQFz5gZ3iSXMVhK0S
   - TikTok: https://docs.google.com/spreadsheets/d/1k22c3PGY6aQjygAX6df_rQLR8aTzL-iz
   - Lazada: https://docs.google.com/spreadsheets/d/1FxLAUiwabmNcBc3TA-bpHqg2MSK7uJ4U
   - TikTok Content: https://docs.google.com/spreadsheets/d/1OWZGQD1wHvIlLAAg_7rJ0CtL-X9vEx4py-cRp_e3NvQ
===========================================
"""

import os, re, json, sys, glob, csv, io, logging, datetime, subprocess
from collections import defaultdict
from pathlib import Path

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE   = Path('/Users/thanasablilutanon/Library/CloudStorage/'
              'GoogleDrive-thanasab.li@gmail.com/.shortcut-targets-by-id/'
              '1-TeohYqk3oWyyTHTbnLIjXW8mAqYowRe/Digital Marketing/claude/All')
CREDS  = Path.home() / 'wibwub-service-account.json'
LOG_FILE = Path.home() / 'Library/Logs/wibwub_update.log'

# ─── Logging ─────────────────────────────────────────────────────────────────
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    filename=str(LOG_FILE), level=logging.INFO,
    format='[%(asctime)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S'
)
def log(msg):  logging.info(msg);  print(msg)
def err(msg):  logging.error(msg); print(f'ERROR: {msg}', file=sys.stderr)

# ─── Google Sheets IDs / gids ─────────────────────────────────────────────────
SHEETS = {
    'shopee':  ('10LrzWB8bbCO9FigCQFz5gZ3iSXMVhK0S', 1820466351),
    'lazada':  ('1FxLAUiwabmNcBc3TA-bpHqg2MSK7uJ4U', 1032656124),
    'tiktok':  ('1k22c3PGY6aQjygAX6df_rQLR8aTzL-iz',  150856480),
    'content': ('1OWZGQD1wHvIlLAAg_7rJ0CtL-X9vEx4py-cRp_e3NvQ', 1242793042),
}

# ═════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═════════════════════════════════════════════════════════════════════════════
def parse_num(s):
    if not s: return 0
    try: return float(re.sub(r'[฿,\s%]', '', str(s)))
    except: return 0

def replace_arr(html, name, values, qs=False):
    vs = (json.dumps(values, ensure_ascii=False) if qs
          else '[' + ','.join(str(v) for v in values) + ']')
    # ใช้ lambda เป็น repl เพื่อไม่ให้ re.sub พยายาม decode backslash-escape
    # ในข้อมูลจริง (เช่น \u, \p ที่มาจากชื่อ/ข้อความ) จนพัง "bad escape"
    return re.sub(rf'(const {re.escape(name)}\s*=\s*)\[.*?\]',
                  lambda m: m.group(1) + vs, html, flags=re.DOTALL)

def extract_arr(html, name):
    """อ่านค่าปัจจุบันของ const array จาก HTML — ใช้เทียบกันก่อนเขียนทับ
    เพื่อกันกรณี fetch ข้อมูลใหม่ล้มเหลวบางส่วนแล้วเขียน 0 ทับข้อมูลจริงที่มีอยู่แล้ว"""
    m = re.search(rf'const {re.escape(name)}\s*=\s*(\[.*?\])', html, re.DOTALL)
    if not m:
        return None
    try:
        return json.loads(m.group(1))
    except Exception:
        return None

def write_if_changed(path, html, label):
    path = Path(path)
    if not path.exists(): err(f'{label}: file not found'); return False
    orig = path.read_text(encoding='utf-8')
    if html == orig: log(f'  ℹ️  {label}: no change'); return False
    path.write_text(html, encoding='utf-8')
    log(f'  ✅ {label}: updated'); return True

# ═════════════════════════════════════════════════════════════════════════════
# STEP 0: ตรวจ Google Drive
# ═════════════════════════════════════════════════════════════════════════════
def check_base():
    if not BASE.exists():
        err(f'Google Drive ไม่ได้ mount: {BASE}')
        sys.exit(1)
    log(f'[STEP 0] BASE: {BASE}')

# ═════════════════════════════════════════════════════════════════════════════
# STEP 1: อ่าน Google Sheets ผ่าน gspread
# ═════════════════════════════════════════════════════════════════════════════
def read_sheets():
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        from google.auth.transport.requests import AuthorizedSession
    except ImportError:
        err('กรุณาติดตั้ง: pip3 install gspread google-auth --break-system-packages')
        sys.exit(1)

    if not CREDS.exists():
        err(f'ไม่พบ service account: {CREDS}')
        err('ดูคำแนะนำใน README ด้านบนของไฟล์นี้')
        sys.exit(1)

    scopes = [
        'https://www.googleapis.com/auth/spreadsheets.readonly',
        'https://www.googleapis.com/auth/drive.readonly',
    ]
    creds   = Credentials.from_service_account_file(str(CREDS), scopes=scopes)
    gc      = gspread.authorize(creds)
    session = AuthorizedSession(creds)   # ใช้ download xlsx ผ่าน Drive API

    def fetch_via_drive(sheet_id, gid):
        """ดาวน์โหลด sheet ผ่าน Drive export (รองรับ .xlsx)"""
        # ลอง CSV export สำหรับ tab ที่ระบุก่อน
        url = (f'https://docs.google.com/spreadsheets/d/{sheet_id}'
               f'/export?format=csv&gid={gid}')
        resp = session.get(url)
        if resp.status_code == 200:
            return list(csv.reader(io.StringIO(resp.text, newline='')))

        # fallback: export ทั้งไฟล์เป็น xlsx → หา sheet ที่มี sales data
        try:
            import openpyxl
        except ImportError:
            resp.raise_for_status()   # re-raise original error

        url_xlsx = (f'https://docs.google.com/spreadsheets/d/{sheet_id}'
                    f'/export?format=xlsx')
        r2 = session.get(url_xlsx)
        r2.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(r2.content),
                                    data_only=True, read_only=True)
        # หา sheet ที่มี date pattern "01-XX/MM/YY" = sales data
        target = None
        for ws in wb.worksheets:
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 30: break
                if row and re.match(r'01-\d{2}/\d{2}/\d{2}', str(row[0] or '')):
                    target = ws; break
            if target: break
        if not target:
            target = wb.worksheets[0]
        rows = [[str(c) if c is not None else '' for c in row]
                for row in target.iter_rows(values_only=True)]
        wb.close()
        return rows

    data = {}
    log('[STEP 1] อ่าน Google Sheets...')
    for name, (sheet_id, gid) in SHEETS.items():
        try:
            sh = gc.open_by_key(sheet_id)
            ws = next((w for w in sh.worksheets() if w.id == gid), sh.sheet1)
            data[name] = ws.get_all_values()
            log(f'  {name}: {len(data[name])} rows')
        except Exception as e:
            if '400' in str(e) or 'Office file' in str(e):
                # ไฟล์เป็น .xlsx (ไม่ใช่ native Google Sheet) — ใช้ Drive export แทน
                try:
                    data[name] = fetch_via_drive(sheet_id, gid)
                    log(f'  {name}: {len(data[name])} rows (Drive export)')
                except Exception as e2:
                    err(f'  {name}: Drive export failed: {e2}')
                    data[name] = []
            else:
                err(f'  {name}: {e}')
                data[name] = []
    return data

# ═════════════════════════════════════════════════════════════════════════════
# STEP 2: Process Sales (Shopee / TikTok / Lazada)
# ═════════════════════════════════════════════════════════════════════════════
def process_sales(sheets):
    log('[STEP 2] Process Sales...')
    ALL_MONTHS = ['01/25','02/25','03/25','04/25','05/25','06/25','07/25','08/25',
                  '09/25','10/25','11/25','12/25',
                  '01/26','02/26','03/26','04/26','05/26','06/26','07/26',
                  '01/27','02/27','03/27','04/27','05/27','06/27','07/27',
                  '08/27','09/27','10/27','11/27','12/27']
    M_LBL = {'01':'ม.ค.','02':'ก.พ.','03':'มี.ค.','04':'เม.ย.','05':'พ.ค.',
              '06':'มิ.ย.','07':'ก.ค.','08':'ส.ค.','09':'ก.ย.',
              '10':'ต.ค.','11':'พ.ย.','12':'ธ.ค.'}

    # M5 ใน Dashboard ใช้ชื่อเดือนเท่านั้น (ไม่มีปี) เช่น ['ม.ค.','ก.พ.',...,'พ.ค.']
    def to_label(key):
        mm, yy = key.split('/')
        return M_LBL.get(mm, mm)  # เดือนอย่างเดียว ไม่ใส่ปี

    def extract(rows):
        """หา row สุดท้ายของแต่ละเดือน (01-28..31/MM/YY)"""
        monthly = {}
        for row in rows:
            if not row: continue
            d = str(row[0]).strip()
            # end-of-month: "01-31/05/26"
            m = re.match(r'01-(?:2[89]|3[01])/(\d{2})/(\d{2})$', d)
            if not m:
                # partial month that hasn't reached end yet: "01-17/05/26"
                m = re.match(r'01-\d{2}/(\d{2})/(\d{2})$', d)
                if not m: continue
            key = f'{m.group(1)}/{m.group(2)}'
            # always overwrite → สุดท้ายที่พบ = ยอดสะสมล่าสุด
            def g(i): return parse_num(row[i]) if i < len(row) else 0
            monthly[key] = row   # เก็บ raw row ไว้ parse ที่ caller
        return monthly

    def parse_shopee(rows):
        # tab "ยอดรายเดือน": 1=ยอดขาย 2=ads 3=ค่าธรรมเนียม 4=คูปอง 5=%คูปอง
        #   6=ลูกค้าใหม่ 7=ลูกค้าเก่า 8=จำนวน order 9=ยกเลิก 10=%ยกเลิก
        raw = extract(rows)
        out = {}
        for key, row in raw.items():
            def g(i): return parse_num(row[i]) if i < len(row) else 0
            out[key] = {'rev': g(1), 'ads': g(2), 'fee': g(3),
                        'cancel_pct': g(10), 'ord': g(8), 'cancel': g(9)}
        return out

    def parse_tiktok(rows):
        # tab "ยอดรายเดือน": 1-3=ยอด Afi (อย่าใช้), 5=ยอดขายรวม 8=ads 9=ค่าธรรมเนียม
        #   16=จำนวน order 17=ยกเลิก 18=%ยกเลิก
        raw = extract(rows)
        out = {}
        for key, row in raw.items():
            def g(i): return parse_num(row[i]) if i < len(row) else 0
            out[key] = {'rev': g(5), 'ads': g(8), 'fee': g(9),
                        'cancel_pct': g(18), 'ord': g(16), 'cancel': g(17)}
        return out

    def parse_lazada(rows):
        # tab "ยอดรายเดือน": 1=ยอดขาย 2=ads 3=ค่าธรรมเนียม 5=% 8=จำนวน order
        raw = extract(rows)
        out = {}
        for key, row in raw.items():
            def g(i): return parse_num(row[i]) if i < len(row) else 0
            out[key] = {'rev': g(1), 'ads': g(2), 'fee': g(3),
                        'cost_pct': g(5), 'ord': g(8)}
        return out

    shopee = parse_shopee(sheets.get('shopee', []))
    tiktok = parse_tiktok(sheets.get('tiktok', []))
    lazada = parse_lazada(sheets.get('lazada', []))

    # กรองเฉพาะเดือนของปีปัจจุบัน (เช่น XX/26) เพื่อให้ตรงกับ M5 ใน Dashboard
    # M5 = ['ม.ค.','ก.พ.',...] ใช้ชื่อเดือนเท่านั้น ไม่มีปี → ต้องมีเดือนเท่ากัน
    # ใช้ union ของทั้ง 3 แพลตฟอร์ม (ไม่ใช่แค่ shopee) — เดิมถ้า shopee ยังไม่มีเดือนนั้น
    # (แต่ lazada/tiktok มีแล้ว) เดือนนั้นจะหายไปทั้งเดือนจาก Dashboard ทั้งที่มีข้อมูลบางส่วน
    current_yr = datetime.datetime.now().strftime('%y')  # '26'
    months_seen = set(shopee) | set(tiktok) | set(lazada)
    available = [m for m in ALL_MONTHS if m in months_seen and m.endswith(f'/{current_yr}')]
    if not available:
        err('  ไม่พบข้อมูล Sales (shopee/tiktok/lazada ว่างทั้งหมด) — ข้าม Sales'); return None
    missing = [(m, [nm for nm, d in (('shopee',shopee),('tiktok',tiktok),('lazada',lazada)) if m not in d])
               for m in available]
    for m, plats in missing:
        if plats:
            log(f'  WARNING: {m} ไม่มีข้อมูลจาก {", ".join(plats)} (ใช้ 0 แทน)')

    labels = [to_label(m) for m in available]
    n = len(available)
    log(f'  Sales: {n} months, latest={labels[-1]}')

    def sh(m, k): return shopee.get(m, {}).get(k, 0)
    def tk(m, k): return tiktok.get(m, {}).get(k, 0)
    def lz(m, k): return lazada.get(m, {}).get(k, 0)

    return {
        'M4': labels,
        'SH_REV':         [int(sh(m,'rev'))         for m in available],
        'SH_ORD':         [int(sh(m,'ord'))          for m in available],
        'SH_CANCEL_PCT':  [round(sh(m,'cancel_pct'),2) for m in available],
        'TK_REV':         [round(tk(m,'rev'),2)      for m in available],
        'TK_ORD':         [int(tk(m,'ord'))           for m in available],
        'TK_CANCEL_PCT':  [round(tk(m,'cancel_pct'),2) for m in available],
        'TK_ADSSPEND':    [int(tk(m,'ads'))           for m in available],
        'TK_FEECOMM':     [int(tk(m,'fee'))           for m in available],
        'LZ_REV':         [int(lz(m,'rev'))           for m in available],
        'LZ_ORD':         [int(lz(m,'ord'))           for m in available],
        'LZ_COST_PCT':    [round(lz(m,'cost_pct'),2) for m in available],
        'n_months': n, 'latest_month': labels[-1],
    }

# ═════════════════════════════════════════════════════════════════════════════
# STEP 3: Process Affiliate (จาก Data Affiliate/ xlsx)
# ═════════════════════════════════════════════════════════════════════════════
def process_affiliate():
    log('[STEP 3] Process Affiliate (xlsx)...')
    try:
        import pandas as pd
    except ImportError:
        err('กรุณาติดตั้ง: pip3 install pandas openpyxl --break-system-packages')
        return None

    AFI_DIR = BASE / 'Data Affiliate'
    if not AFI_DIR.exists():
        err(f'  Data Affiliate/ ไม่พบ: {AFI_DIR}'); return None

    # เดือนเริ่มต้น = พย.68 (พ.ย. 2025) = index 0 — คำนวณเดือนอัตโนมัติ ไม่ต้องแก้ทุกเดือน
    TH_ABBR = {1:'มค',2:'กพ',3:'มีนา',4:'เมษา',5:'พค',6:'มิย',
               7:'กค',8:'สค',9:'กย',10:'ตค',11:'พย',12:'ธค'}
    def ym_to_idx(yr, mo):
        return (yr - 2025) * 12 + (mo - 11)
    def idx_to_label(i):
        mo = (10 + i) % 12 + 1
        yr = 2025 + (10 + i) // 12
        return f'{TH_ABBR[mo]}.{(yr + 543) % 100:02d}'

    def detect_mi(fname):
        # อ่านเดือนจากวันที่ในชื่อไฟล์ (YYYYMMDD...) — รองรับทุกเดือนอัตโนมัติ
        m = re.search(r'(20\d{2})(\d{2})\d{2}', fname)
        if m:
            i = ym_to_idx(int(m.group(1)), int(m.group(2)))
            return i if i >= 0 else None
        # fallback: ชื่อเดือนภาษาอังกฤษ (เดาปีจากไฟล์ date ไม่ได้ จึงข้าม)
        return None

    # สแกนแบบ recursive (os.walk) เพราะไฟล์ Creator_List ฉบับสมบูรณ์ของบางเดือน
    # อาจถูกย้ายไปเก็บในโฟลเดอร์ย่อย (เช่น Data Affiliate/ครีเอเตอร์/) ไม่ใช่แค่ top-level
    # อ่านเฉพาะไฟล์ Creator_List เท่านั้น — ไฟล์ Video_List/Product_List/Live_List/
    # Core_Metrics ใช้ column layout ต่างกัน (col1 อาจเป็น video ID/product ID แทน GMV)
    # ถ้าอ่านปนกัน จะได้ยอด GMV บวก ID ตัวเลขยาวๆ เข้าไปด้วย ทำให้ยอดพังเป็นเลขมหาศาล
    all_files = []
    for root, _dirs, files in os.walk(AFI_DIR):
        for f in files:
            if f.endswith('.xlsx') and not f.startswith(('~', '.')) and 'Creator_List' in f:
                all_files.append(str(Path(root) / f))

    def month_bounds(i):
        mo = (10 + i) % 12 + 1
        yr = 2025 + (10 + i) // 12
        first = datetime.date(yr, mo, 1)
        last = (datetime.date(yr + (mo // 12), mo % 12 + 1, 1) - datetime.timedelta(days=1))
        return first, last

    def parse_range(fname):
        # หาช่วงวันที่ในชื่อไฟล์ "YYYYMMDD-YYYYMMDD"
        m = re.search(r'(20\d{6})-(20\d{6})', fname)
        if not m: return None
        try:
            s = datetime.datetime.strptime(m.group(1), '%Y%m%d').date()
            e = datetime.datetime.strptime(m.group(2), '%Y%m%d').date()
            return s, e
        except Exception:
            return None

    # เลือก "1 ไฟล์ต่อเดือน" คือไฟล์ที่ช่วงวันที่ครอบคลุม (overlap) เดือนนั้นมากที่สุด
    # (ไม่ใช่แค่ end-date ล่าสุด — เพราะไฟล์สัปดาห์สุดท้ายเช่น 20260625-20260701 จะมี
    # end-date ใหม่กว่าไฟล์เต็มเดือน 20260601-20260630 ทั้งที่คลุมข้อมูลเดือนนั้นน้อยกว่ามาก)
    # แทนที่จะไล่ประมวลผลทุกไฟล์ของเดือนเดียวกันแล้ว overwrite ทีละครีเอเตอร์
    # — วิธีเดิมทำให้ครีเอเตอร์ที่หายไปจากไฟล์ใหม่กว่า (แต่มีอยู่ในไฟล์เก่า) ค้างค่าผิดๆ ไว้
    best_per_month = {}   # mi -> (fp, overlap_days, end_date)
    for fp in all_files:
        fname = os.path.basename(fp)
        mi = detect_mi(fname)
        if mi is None: continue
        rng = parse_range(fname)
        if not rng: continue
        s, e = rng
        tfirst, tlast = month_bounds(mi)
        overlap = (min(e, tlast) - max(s, tfirst)).days + 1
        if overlap <= 0: continue
        cur = best_per_month.get(mi)
        if cur is None or (overlap, e) > (cur[1], cur[2]):
            best_per_month[mi] = (fp, overlap, e)

    xlsx_files = sorted(v[0] for v in best_per_month.values())

    # pass 1: หาเดือนล่าสุดเพื่อกำหนดจำนวนเดือน (อย่างน้อยถึง พค.69 = index 6)
    max_idx = 6
    for f in xlsx_files:
        mi = detect_mi(os.path.basename(f))
        if mi is not None:
            max_idx = max(max_idx, mi)
    MONTH_ORDER = [idx_to_label(i) for i in range(max_idx + 1)]

    c_gmv  = defaultdict(lambda: [None]*len(MONTH_ORDER))
    c_net  = defaultdict(lambda: [None]*len(MONTH_ORDER))
    c_comm = defaultdict(lambda: [None]*len(MONTH_ORDER))
    c_ord  = defaultdict(lambda: [None]*len(MONTH_ORDER))
    c_ret  = defaultdict(lambda: [None]*len(MONTH_ORDER))

    # เก็บช่วงวัน (start_day-end_day) ของไฟล์ที่เลือกใช้ต่อเดือน — ใช้ทำ label
    # แบบ "มิ.ย. (1-30)" / "ก.ค. (1-5)" ให้ตรงกับ format เดิมที่เคยแก้ด้วยมือ
    day_range = {}
    for fp in xlsx_files:
        fname = os.path.basename(fp)
        mi = detect_mi(fname)
        if mi is not None:
            rng = parse_range(fname)
            if rng:
                s, e = rng
                tfirst, tlast = month_bounds(mi)
                ds = max(s, tfirst).day
                de = min(e, tlast).day
                # หมายเหตุ: แสดง (start-end) ทุกเดือนที่มีไฟล์จริง (ทั้งเดือนเต็ม/ไม่เต็ม)
                # เพื่อความชัดเจนของช่วงข้อมูล — ต่างจาก format เดิมที่เคยแก้ด้วยมือซึ่งซ่อน
                # suffix นี้ไว้เฉพาะเดือนเก่าๆ อย่างไม่สม่ำเสมอ (บาง full month มี บางอันไม่มี)
                day_range[mi] = (ds, de)

    for fp in xlsx_files:
        fname = os.path.basename(fp)
        mi = detect_mi(fname)
        if mi is None: log(f'  WARNING: ไม่รู้เดือนของ {fname}'); continue
        try:
            df = pd.read_excel(fp, header=None)
            cnt = 0
            for _, row in df.iterrows():
                name = str(row.iloc[0]).strip().lstrip('@') if pd.notna(row.iloc[0]) else ''
                if not name or name in ['nan','Creator','ชื่อครีเอเตอร์'] or 'Unnamed' in name: continue
                try:
                    def pn(v): return parse_num(v) if pd.notna(v) else 0
                    gmv  = pn(row.iloc[1])  if len(row) > 1  else 0
                    ret  = pn(row.iloc[2])  if len(row) > 2  else 0
                    ordn = pn(row.iloc[3])  if len(row) > 3  else 0
                    comm = pn(row.iloc[10]) if len(row) > 10 else 0
                    # sanity guard: ยอด GMV ต่อครีเอเตอร์ต่อเดือนจริงไม่เคยเกินหลักล้านบาท
                    # ถ้าเจอเลขมหาศาล (เช่น video ID/product ID หลุดเข้ามา) ให้ข้ามแถวนี้ทิ้ง
                    if gmv > 10_000_000 or comm > 10_000_000:
                        log(f'  WARNING: ข้ามแถวผิดปกติใน {fname} name={name!r} gmv={gmv} comm={comm}')
                        continue
                    if gmv > 0:
                        c_gmv[name][mi]  = int(gmv)
                        c_net[name][mi]  = int(gmv - ret)
                        c_comm[name][mi] = int(comm)
                        c_ord[name][mi]  = int(ordn)
                        c_ret[name][mi]  = int(ret)
                        cnt += 1
                except: pass
            log(f'  {fname} → {MONTH_ORDER[mi]}: {cnt} creators')
        except Exception as e:
            err(f'  ERROR {fname}: {e}')

    creators = []
    for name in set(c_gmv.keys()):
        mn = c_gmv[name]
        total = sum(v for v in mn if v)
        if not total: continue
        tot_ret  = sum(v for v in c_ret[name]  if v)
        tot_comm = sum(v for v in c_comm[name] if v)
        tot_ord  = sum(v for v in c_ord[name]  if v)
        creators.append({
            'n': name, 't': total, 'ma': sum(1 for v in mn if v), 'mn': mn,
            'returns': tot_ret, 'net': total - tot_ret, 'comm': tot_comm,
            'orders': tot_ord,
            'comm_rate': round(tot_comm / total * 100, 1) if total else 0,
        })
    creators.sort(key=lambda x: -x['t'])

    gmvD  = [sum(c_gmv[c['n']][i]  or 0 for c in creators) for i in range(len(MONTH_ORDER))]
    netD  = [sum(c_net[c['n']][i]  or 0 for c in creators) for i in range(len(MONTH_ORDER))]
    commD = [sum(c_comm[c['n']][i] or 0 for c in creators) for i in range(len(MONTH_ORDER))]
    crD   = [sum(1 for c in creators if (c_gmv[c['n']][i] or 0) > 0) for i in range(len(MONTH_ORDER))]

    log(f'  Affiliate: {len(creators)} creators')
    log(f'  gmvD ={gmvD}')
    log(f'  netD ={netD}   ← ตรวจว่า net ≤ gmv ทุกเดือน')
    log(f'  commD={commD}')

    # ── สร้างชุดข้อมูลพร้อมเขียนลง HTML โดยตรง (ตรงกับชื่อตัวแปรจริงใน
    #    WIBWUB_Affiliate_Dashboard.html: AF_MO/AF_GMV/AF_NET/AF_COM/AF_CR
    #    (เริ่มนับที่ ม.ค. ของปีปัจจุบัน) และ CREATORS/CREATOR_MONTHS
    #    (เริ่มนับที่ มี.ค. ของปีปัจจุบัน) — ป้องกันบัคเดิมที่สคริปต์เขียนไปยัง
    #    ชื่อตัวแปรเก่า (gmvD/netD/commD/crD/months/raw) ที่ไม่มีอยู่จริงในไฟล์
    #    ทำให้ re.sub ไม่เจอ match แล้วไม่อัปเดตอะไรเลยแบบเงียบๆ ทุกเดือน
    THAI_M = {1:'ม.ค.',2:'ก.พ.',3:'มี.ค.',4:'เม.ย.',5:'พ.ค.',6:'มิ.ย.',
              7:'ก.ค.',8:'ส.ค.',9:'ก.ย.',10:'ต.ค.',11:'พ.ย.',12:'ธ.ค.'}
    cur_year = datetime.datetime.now().year
    jan_idx = ym_to_idx(cur_year, 1)
    mar_idx = ym_to_idx(cur_year, 3)

    af_mo = []
    for i in range(max(jan_idx, 0), len(MONTH_ORDER)):
        mo = (10 + i) % 12 + 1
        lbl = THAI_M[mo]
        if i in day_range:
            lbl += f' ({day_range[i][0]}-{day_range[i][1]})'
        af_mo.append(lbl)
    af_gmv = gmvD[max(jan_idx, 0):]
    af_net = netD[max(jan_idx, 0):]
    af_com = commD[max(jan_idx, 0):]
    af_cr  = crD[max(jan_idx, 0):]

    creators_out = [{
        'name': c['n'], 'gmv': c['t'], 'returns': c['returns'], 'net': c['net'],
        'comm': c['comm'], 'orders': c['orders'], 'comm_rate': c['comm_rate'],
    } for c in creators if c['t'] >= 1000]

    creator_months_out = {
        c['n']: [v or 0 for v in c['mn'][max(mar_idx, 0):]] for c in creators
    }

    return {'months': MONTH_ORDER, 'gmvD': gmvD, 'netD': netD,
            'commD': commD, 'crD': crD, 'creators': creators,
            'AF_MO': af_mo, 'AF_GMV': af_gmv, 'AF_NET': af_net,
            'AF_COM': af_com, 'AF_CR': af_cr,
            'CREATORS': creators_out, 'CREATOR_MONTHS': creator_months_out}

# ═════════════════════════════════════════════════════════════════════════════
# STEP 4: Process TikTok Content
# ═════════════════════════════════════════════════════════════════════════════
def process_tiktok_content(rows):
    log('[STEP 4] Process TikTok Content...')
    hdr = None; data_start = 0
    for i, row in enumerate(rows):
        joined = ' '.join(row).lower()
        if ('view' in joined or 'วิว' in joined) and ('pillar' in joined or 'content' in joined):
            hdr = row; data_start = i+1; break
    if not hdr: log('  WARNING: header ไม่พบใน TikTok Content sheet'); return []

    def ci(kws):
        for j, h in enumerate(hdr):
            if any(k in h.lower().strip() for k in kws): return j
        return None

    c_lbl=ci(['วันที่','lbl','date']); c_day=ci(['day','dow'])
    c_mo=ci(['month','เดือน']); c_pil=ci(['pillar','category'])
    c_con=ci(['content','caption']); c_v=ci(['view','วิว'])
    c_eng=ci(['eng','engagement']); c_love=ci(['love','like'])
    c_com=ci(['comment']); c_sha=ci(['share']); c_sav=ci(['save'])
    c_ret=ci(['ret','retention']); c_wat=ci(['watch'])
    c_er=ci(['er','rate']); c_url=ci(['url','tiktok.com'])

    def gv(row, c): return row[c] if c is not None and c < len(row) else ''

    posts = []
    for row in rows[data_start:]:
        if len(row) < 3: continue
        try:
            v = int(parse_num(gv(row, c_v)))
            lbl = gv(row, c_lbl).strip()
            if not v and not lbl: continue
            def jv(v): return "'"+str(v).replace("'","\\'")+  "'" if isinstance(v,str) else v
            posts.append({
                'lbl':  lbl,
                'day':  int(gv(row,c_day)) if str(gv(row,c_day)).strip().isdigit() else 0,
                'month':int(gv(row,c_mo))  if str(gv(row,c_mo)).strip().isdigit()  else 0,
                'pillar':  gv(row,c_pil).strip(),
                'content': gv(row,c_con).strip(),
                'views':   v,
                'eng':     int(parse_num(gv(row,c_eng))),
                'love':    int(parse_num(gv(row,c_love))),
                'comment': int(parse_num(gv(row,c_com))),
                'share':   int(parse_num(gv(row,c_sha))),
                'save':    int(parse_num(gv(row,c_sav))),
                'ret':     float(parse_num(gv(row,c_ret))),
                'watch':   float(parse_num(gv(row,c_wat))),
                'er':      float(parse_num(gv(row,c_er))),
                'url':     gv(row,c_url).strip() or '#',
            })
        except: pass
    log(f'  TikTok Content: {len(posts)} posts')
    return posts

# ═════════════════════════════════════════════════════════════════════════════
# STEP 5: Process Shipnity (Sales_Dashboard)
# ═════════════════════════════════════════════════════════════════════════════
def process_shipnity():
    log('[STEP 5] Process Shipnity...')
    try:
        import openpyxl
    except ImportError:
        err('กรุณาติดตั้ง: pip3 install openpyxl --break-system-packages'); return None

    SHIP_DIR = BASE / 'Data Shipnity'
    if not SHIP_DIR.exists(): err(f'  Data Shipnity/ ไม่พบ'); return None

    def parse_date(v):
        if v is None: return None
        if isinstance(v, (datetime.datetime, datetime.date)):
            dt = v if isinstance(v, datetime.datetime) else datetime.datetime.combine(v, datetime.time())
            return dt.strftime('%Y-%m-%d')
        s = str(v).strip()
        for fmt in ['%d/%m/%Y %H:%M','%d/%m/%Y','%Y-%m-%d %H:%M:%S','%Y-%m-%d']:
            try: return datetime.datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except: pass
        return None

    CHANNEL_MAP = {'tiktok':'TikTok','tiktokshop':'TikTok','shopee':'Shopee',
                   'facebook':'Facebook','fb':'Facebook','lazada':'Lazada',
                   'line oa':'Line OA','line shopping':'Line Shopping',
                   'carcare':'Carcare','marketing':'Marketing','pos':'POS',
                   'website':'Website','เบิกของ':'เบิกของ'}
    CHANNELS = ["Carcare","Facebook","Lazada","Line OA","Line Shopping",
                "Marketing","POS","Shopee","TikTok","Website","เบิกของ"]
    CH_IDX = {c.lower(): i for i, c in enumerate(CHANNELS)}

    def norm_ch(raw):
        if not raw: return 'Other'
        r = str(raw).lower().strip()
        return CHANNEL_MAP.get(r, raw.strip())

    files = [f for f in glob.glob(str(SHIP_DIR / '*.xlsx'))
             if not os.path.basename(f).startswith('~')]
    agg = defaultdict(lambda: {'q':0, 's':0.0})
    product_set = set()
    order_seen = {}        # order_id → {d, ch, s} — ใช้ dedup order counts
    seen_order_prod = set()  # (order_id, name) — dedup ไฟล์ที่ซ้อนทับกัน

    for fp in sorted(files):
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue
            if not row or all(c is None for c in row[:6]): continue
            date_val = None
            for di in [19, 12, 11]:
                if di < len(row) and row[di]:
                    d = parse_date(row[di])
                    if d: date_val = d; break
            if not date_val: continue
            name = str(row[1]).strip() if row[1] else ''
            if not name: continue
            unit_price  = parse_num(row[2]) if len(row) > 2 and row[2] else 0
            qty         = int(parse_num(row[3])) if len(row) > 3 and row[3] else 1
            rev         = unit_price * max(1, qty)
            order_id    = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            order_total = parse_num(row[5]) if len(row) > 5 and row[5] else 0
            ch          = norm_ch(str(row[15]).strip() if len(row) > 15 and row[15] else '')
            if rev <= 0: continue
            # FIX: dedup (order_id, product) เพื่อป้องกัน double-count จากไฟล์ที่ซ้อนทับกัน
            op_key = (order_id, name) if order_id else None
            if op_key and op_key in seen_order_prod: continue
            if op_key: seen_order_prod.add(op_key)
            agg[(date_val, name, ch)]['q'] += qty
            agg[(date_val, name, ch)]['s'] += rev
            product_set.add(name)
            # ติดตาม unique orders (dedup ด้วย order_id) สำหรับ order count + AOV
            if order_id and order_id not in order_seen and order_total > 0:
                order_seen[order_id] = {'d': date_val, 'ch': ch, 's': order_total}
        wb.close()

    products = sorted(product_set)
    PROD_IDX = {p: i for i, p in enumerate(products)}
    lines = []
    for (date, prod, ch), vals in sorted(agg.items()):
        p_i = PROD_IDX.get(prod)
        ch_l = ch.lower()
        c_i  = CH_IDX.get(ch_l)
        if c_i is None:
            c_i = len(CHANNELS); CHANNELS.append(ch); CH_IDX[ch_l] = c_i
        if p_i is None: continue
        lines.append({'d': date, 'p': p_i, 'c': c_i,
                      'q': vals['q'], 's': round(vals['s'], 2)})
    lines.sort(key=lambda x: x['d'])

    # สร้าง order_counts: [{d, c, n, r}] เพื่อให้ JS คำนวณ total_orders และ AOV ได้ถูกต้อง
    oc_agg = {}
    for oid, info in order_seen.items():
        ch_l = info['ch'].lower()
        c_i  = CH_IDX.get(ch_l)
        if c_i is None:
            c_i = len(CHANNELS); CHANNELS.append(info['ch']); CH_IDX[ch_l] = c_i
        key = (info['d'], c_i)
        if key not in oc_agg:
            oc_agg[key] = {'n': 0, 'r': 0.0}
        oc_agg[key]['n'] += 1
        oc_agg[key]['r'] += info['s']
    order_counts = [{'d': d, 'c': c, 'n': v['n'], 'r': round(v['r'], 2)}
                    for (d, c), v in sorted(oc_agg.items())]

    raw = {'date_min': lines[0]['d'] if lines else '',
           'date_max': lines[-1]['d'] if lines else '',
           'channels': CHANNELS, 'products': products,
           'lines': lines, 'order_counts': order_counts}
    log(f'  Shipnity: {len(lines)} lines, {len(order_counts)} order-day-ch groups, {raw["date_min"]} → {raw["date_max"]}')
    return raw

# ═════════════════════════════════════════════════════════════════════════════
# STEP 6: Update HTML Files
# ═════════════════════════════════════════════════════════════════════════════
def update_html(sales, aff, posts, shipnity):
    log('[STEP 6] Update HTML files...')
    # TK_CANCEL_PCT ถูกตัดออก — คอลัมน์ 5 ของ TikTok sheet คือ Gross GMV ไม่ใช่ cancel rate
    SALES_ARRS = ['SH_REV','SH_ORD','SH_CANCEL_PCT','TK_REV','TK_ORD',
                  'TK_ADSSPEND','TK_FEECOMM','LZ_REV','LZ_ORD','LZ_COST_PCT']

    # A. Dashboard + Mobile — อัปเดต Sales arrays พร้อมกัน
    if sales:
        THAI_M = {1:'ม.ค.',2:'ก.พ.',3:'มี.ค.',4:'เม.ย.',5:'พ.ค.',6:'มิ.ย.',
                  7:'ก.ค.',8:'ส.ค.',9:'ก.ย.',10:'ต.ค.',11:'พ.ย.',12:'ธ.ค.'}
        _now = datetime.datetime.now()
        badge_str = f'อัปเดต {_now.day} {THAI_M[_now.month]} {_now.year + 543}'
        for fname in ['WIBWUB_Dashboard.html', 'WIBWUB_Mobile.html']:
            fp = BASE / fname
            if not fp.exists(): log(f'  SKIP {fname}'); continue
            html = fp.read_text(encoding='utf-8')
            # M5 คือตัวแปร month labels ใน Dashboard/Mobile (ไม่ใช่ M4)
            if 'M5' in html:
                html = replace_arr(html, 'M5', sales['M4'], qs=True)
            for k in SALES_ARRS:
                if k in html:
                    new_vals = sales[k]
                    old_vals = extract_arr(html, k)
                    if old_vals and len(old_vals) == len(new_vals):
                        merged = []
                        for i, nv in enumerate(new_vals):
                            ov = old_vals[i]
                            # กัน fetch fail บางส่วน (เช่น Sheet เข้าไม่ได้ชั่วคราว) เขียน 0
                            # ทับข้อมูลจริงที่มีอยู่แล้ว — ถ้าค่าใหม่เป็น 0 แต่ค่าเดิมไม่ใช่ 0/None
                            # ให้คงค่าเดิมไว้แทน (revenue สะสมไม่ควรลดกลับไป 0 กลางเดือน)
                            if nv == 0 and ov not in (0, None):
                                merged.append(ov)
                                log(f'  ⚠️  {k}[{i}] ({sales["M4"][i]}): ได้ 0 จาก fetch แต่ของเดิมคือ {ov} — คงค่าเดิมไว้ (ป้องกันข้อมูลหาย)')
                            else:
                                merged.append(nv)
                        html = replace_arr(html, k, merged)
                    else:
                        html = replace_arr(html, k, new_vals)
            # อัปเดต badge "อัปเดต D MMM YYYY" ใน header ให้ตรงกับวันที่รันจริง
            # (เดิม hardcode ไว้ตอนสร้างไฟล์ ไม่เคยถูกแตะโดยสคริปต์นี้มาก่อน — ทำให้ badge ค้างวันเก่า)
            html = re.sub(
                r'(id="dash-updated"[^>]*>)อัปเดต \d{1,2} \S+ \d{4}',
                rf'\g<1>{badge_str}',
                html
            )
            write_if_changed(fp, html, fname)

    # B. Affiliate Dashboard
    # หมายเหตุ: ตัวแปรจริงในไฟล์คือ AF_MO/AF_GMV/AF_NET/AF_COM/AF_CR (KPI trend)
    # และ CREATORS/CREATOR_MONTHS (ตาราง creator) — ไม่ใช่ gmvD/netD/commD/crD/months/raw
    # ที่เคยเขียนผิดมาก่อน (เป็นสาเหตุที่ข้อมูล affiliate ไม่เคยถูกอัปเดตโดยอัตโนมัติ)
    if aff:
        fp = BASE / 'WIBWUB_Affiliate_Dashboard.html'
        if fp.exists():
            html = fp.read_text(encoding='utf-8')

            # B1. KPI trend arrays (Jan-indexed)
            html = replace_arr(html, 'AF_MO',  aff['AF_MO'],  qs=True)
            html = replace_arr(html, 'AF_GMV', aff['AF_GMV'])
            html = replace_arr(html, 'AF_NET', aff['AF_NET'])
            html = replace_arr(html, 'AF_COM', aff['AF_COM'])
            html = replace_arr(html, 'AF_CR',  aff['AF_CR'])

            # B2. CREATORS (creator totals table)
            def esc(s):
                return str(s).replace('\\', '\\\\').replace('"', '\\"')
            creators_js = ',\n'.join(
                '  {name:"'+esc(c['name'])+'",gmv:'+str(c['gmv'])+',returns:'+str(c['returns'])
                +',net:'+str(c['net'])+',comm:'+str(c['comm'])+',orders:'+str(c['orders'])
                +',comm_rate:'+str(c['comm_rate'])+'}'
                for c in aff['CREATORS']
            )
            creators_block = 'const CREATORS = [\n' + creators_js + '\n];'
            # ใช้ lambda เป็น repl กัน re.sub ตีความ backslash ในชื่อ creator เป็น escape sequence
            html = re.sub(r'const CREATORS\s*=\s*\[[\s\S]*?\];', lambda m: creators_block, html)

            # B3. CREATOR_MONTHS (per-creator per-month GMV, Mar-indexed)
            def esc_key(s):
                return str(s).replace('\\', '\\\\').replace("'", "\\'")
            cm_js = ',\n'.join(
                "  '"+esc_key(name)+"':["+','.join(str(v) for v in months)+']'
                for name, months in aff['CREATOR_MONTHS'].items()
            )
            cm_block = 'const CREATOR_MONTHS = {\n' + cm_js + '\n};'
            html = re.sub(r'const CREATOR_MONTHS\s*=\s*\{[\s\S]*?\};', lambda m: cm_block, html)

            write_if_changed(fp, html, 'WIBWUB_Affiliate_Dashboard.html')

    # B1b. Bump cache-busting version บน iframe ที่ WIBWUB_Dashboard.html ใช้ฝัง Affiliate Dashboard
    # (ป้องกันปัญหา: แก้ WIBWUB_Affiliate_Dashboard.html แล้ว แต่ iframe หลักยังแคชโค้ดเก่าเพราะ ?v= ไม่เคยเปลี่ยน)
    if aff:
        fp = BASE / 'WIBWUB_Dashboard.html'
        if fp.exists():
            html = fp.read_text(encoding='utf-8')
            m = re.search(r'WIBWUB_Affiliate_Dashboard\.html\?v=(\d+)', html)
            if m:
                old_v = int(m.group(1)); new_v = old_v + 1
                html = html.replace(f'WIBWUB_Affiliate_Dashboard.html?v={old_v}',
                                     f'WIBWUB_Affiliate_Dashboard.html?v={new_v}')
                write_if_changed(fp, html, 'WIBWUB_Dashboard.html (affiliate iframe cache-bust)')
                log(f'  ✅ Affiliate iframe cache-bust: v{old_v} → v{new_v}')
            else:
                log('  WARNING: ไม่พบ WIBWUB_Affiliate_Dashboard.html?v=N pattern ใน WIBWUB_Dashboard.html — ข้าม cache-bust')

    # B2. Mobile — AFI arrays (ต้องตรงกับ Affiliate Dashboard เสมอ)
    if aff:
        fp = BASE / 'WIBWUB_Mobile.html'
        if fp.exists():
            html = fp.read_text(encoding='utf-8')
            for k, v in [('AFI_GMV',aff['gmvD']),('AFI_NET',aff['netD']),('AFI_COMM',aff['commD'])]:
                repl_str = f'const {k}=['+','.join(str(x) for x in v)+']'
                html = re.sub(rf'const {k}\s*=\s*\[.*?\]', lambda m, s=repl_str: s, html)
            write_if_changed(fp, html, 'WIBWUB_Mobile.html (AFI arrays)')

    # C. TikTok Dashboard
    if posts:
        def jsv(v):
            if isinstance(v, str): return "'" + v.replace("'", "\\'") + "'"
            return str(v)
        entries = ['  {' + ','.join(f'{k}:{jsv(p[k])}'
                   for k in ['lbl','day','month','pillar','content',
                              'views','eng','love','comment','share','save',
                              'ret','watch','er','url']) + '}'
                   for p in posts]
        new_block = 'ALL_POSTS = [\n' + ',\n'.join(entries) + '\n]'
        fp = BASE / 'WIBWUB_TikTok_Dashboard_v7.html'
        if fp.exists():
            html = fp.read_text(encoding='utf-8')
            html = re.sub(r'ALL_POSTS\s*=\s*\[[\s\S]*?\]', lambda m: new_block, html)
            write_if_changed(fp, html, 'WIBWUB_TikTok_Dashboard_v7.html')

    # D. Sales Dashboard (Shipnity)
    if shipnity:
        fp = BASE / 'Data Shipnity' / 'Sales_Dashboard.html'
        if fp.exists():
            html = fp.read_text(encoding='utf-8')
            raw_js = json.dumps(shipnity, ensure_ascii=False, separators=(',', ':'))
            raw_block = f'const RAW = {raw_js};'
            html = re.sub(r'const RAW\s*=\s*\{[\s\S]*?\};', lambda m: raw_block, html)
            write_if_changed(fp, html, 'Sales_Dashboard.html')

# ═════════════════════════════════════════════════════════════════════════════
# STEP 7: Bump sw.js
# ═════════════════════════════════════════════════════════════════════════════
def bump_sw():
    log('[STEP 7] Bump sw.js...')
    fp = BASE / 'sw.js'
    if not fp.exists(): err(f'  sw.js ไม่พบ'); return
    content = fp.read_text(encoding='utf-8')
    m = re.search(r"const CACHE = 'wibwub-v(\d+)'", content)
    if not m: err('  CACHE pattern ไม่พบ'); return
    old_v, new_v = int(m.group(1)), int(m.group(1)) + 1
    fp.write_text(content.replace(f"'wibwub-v{old_v}'", f"'wibwub-v{new_v}'"), encoding='utf-8')
    log(f'  ✅ sw.js: v{old_v} → v{new_v}')

# ═════════════════════════════════════════════════════════════════════════════
# STEP 8: Git commit + push
# ═════════════════════════════════════════════════════════════════════════════
def git_push():
    log('[STEP 8] Git push...')
    files = ['WIBWUB_Mobile.html', 'WIBWUB_Dashboard.html',
             'WIBWUB_Affiliate_Dashboard.html', 'WIBWUB_TikTok_Dashboard_v7.html',
             'data Ads/WIBWUB_Ads_Dashboard.html',
             'Data Shipnity/Sales_Dashboard.html', 'sw.js',
             'Procurement_Dashboard.html']
    try:
        subprocess.run(['git', 'add'] + files, cwd=str(BASE), check=True,
                       capture_output=True)
        diff = subprocess.run(['git', 'diff', '--cached', '--quiet'], cwd=str(BASE))
        if diff.returncode != 0:
            ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            subprocess.run(['git', 'commit', '-m', f'auto: update {ts}',
                            '--author=WIBWUB Bot <marketingwibwub@gmail.com>'],
                           cwd=str(BASE), check=True, capture_output=True)
        else:
            log('  Git: ไม่มีอะไรเปลี่ยนใหม่ให้ commit')

        # เดิม: ถ้ารอบนี้ไม่มีอะไรใหม่ให้ commit สคริปต์จะ return ทันทีโดยไม่เช็คว่า
        # local มี commit ที่ค้างไม่ได้ push อยู่หรือไม่ (เช่น commit ที่ทำด้วยมือ/
        # ผ่าน Claude นอกรอบ schedule) ทำให้ผู้ใช้ต้องมา push_now.command เองทุกครั้ง
        # แก้โดยเช็ค origin/main..HEAD ทุกรอบ แล้ว push ถ้ามี commit ค้างจริง
        subprocess.run(['git', 'fetch', 'origin', 'main'], cwd=str(BASE),
                       capture_output=True)
        ahead = subprocess.run(['git', 'rev-list', '--count', 'origin/main..HEAD'],
                               cwd=str(BASE), capture_output=True, text=True)
        n_ahead = ahead.stdout.strip() if ahead.returncode == 0 else ''
        if n_ahead and n_ahead != '0':
            r = subprocess.run(['git', 'push', 'origin', 'main'],
                               cwd=str(BASE), capture_output=True, text=True)
            if r.returncode == 0:
                log(f'  ✅ Git pushed ({n_ahead} commit)')
            else:
                err(f'  Git push failed: {r.stderr.strip()}')
        else:
            log('  Git: ไม่มีอะไรต้อง push (up to date กับ origin/main)')
    except subprocess.CalledProcessError as e:
        err(f'  Git error: {e}')
    except Exception as e:
        err(f'  Git error: {e}')

# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    log('=' * 50)
    log('WIBWUB Auto-Update START')
    log('=' * 50)

    check_base()
    sheets    = read_sheets()
    sales     = process_sales(sheets)
    aff       = process_affiliate()
    posts     = process_tiktok_content(sheets.get('content', []))
    shipnity  = process_shipnity()

    update_html(sales, aff, posts, shipnity)
    bump_sw()
    git_push()

    log('=' * 50)
    log('WIBWUB Auto-Update DONE')
    log('=' * 50)
